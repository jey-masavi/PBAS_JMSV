import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from statistics import mean, median, mode, stdev, StatisticsError
import os

# Configuración de rutas
BASE_FOLDER = "PIA"
INPUT_CSV = os.path.join(BASE_FOLDER, "resumen_por_genero.csv")
GRAFICAS_DIR = os.path.join(BASE_FOLDER, "Graficas")
OUTPUT_XLSX = os.path.join(BASE_FOLDER, "visualizacion_spotify.xlsx")

# Crear carpetas si no existen
os.makedirs(GRAFICAS_DIR, exist_ok=True)

# 1. Cargar datos
df = pd.read_csv(INPUT_CSV)

# Validación básica
assert not df.isnull().values.any(), "El CSV contiene valores nulos."
assert all (col in df.columns for col in ['género', 'cantidad_artistas', 'popularidad_media', 'popularidad_máxima', 'popularidad_mínima'])

# 2. Crear gráficas
def guardar_grafica(fig, nombre):
    ruta = os.path.join(GRAFICAS_DIR, nombre)
    fig.savefig(ruta, bbox_inches='tight')
    plt.close(fig)
    return ruta

# 2.1 Gráfico de barras
fig1, ax1 = plt.subplots()
ax1.bar(df['género'], df['cantidad_artistas'], color='skyblue')
ax1.set_title("Cantidad de Artitas por Género")
ax1.set_xlabel("Género")
ax1.set_ylabel("Cantidad de Artistas")
ax1.grid(True, linestyle='--', alpha=0.6)
ruta1 = guardar_grafica(fig1, "barras_artistas.png")

# 2.2 Gráfico de líneas
fig2, ax2 = plt.subplots()
ax2.plot(df['género'], df['popularidad_media'], marker='o', color='green')
ax2.set_title("Popularidad Media por Género")
ax2.set_xlabel("Género")
ax2.set_ylabel("Popularidad Media")
ax2.grid(True, linestyle='--', alpha=0.6)
ruta2 = guardar_grafica(fig2, "linea_popularidad.png")

# 2.3 Diagram de dispersión
fig3, ax3 = plt.subplots()
ax3.scatter(df['popularidad_mínima'], df['popularidad_máxima'], color='purple')
for i, row in df.iterrows():
    ax3.annotate(row['género'], (row['popularidad_mínima'], row['popularidad_máxima']))
ax3.set_title("Dispersión de Popularidad (Min vs Máx)")
ax3.set_xlabel("Popularidad Mínima")
ax3.set_ylabel("Popularidad Máxima")
ax3.grid(True, linestyle='--', alpha=0.6)
ruta3 = guardar_grafica(fig3, "dispersión_popularidad.png")

# 2.4 Gráfico de pastel
fig4, ax4 = plt.subplots()
ax4.pie(df['cantidad_artistas'], labels=df['género'], autopct='%1.1f%%', startangle=140)
ax4.set_title("Proporción de Artistas por Género")
ruta4 = guardar_grafica(fig4, "pastel_proporcion.png")

# 3. Crear archivo Excel
wb = openpyxl.Workbook()

# 3.1 Hoja: Resumen por género
ws1 = wb.active
ws1.title = "Resumen por género"
for r in dataframe_to_rows(df, index=False, header=True):
    ws1.append(r)

# 3.2 Hoja: Estadísticas generales
ws2 = wb.create_sheet("Estadísticas generales")

# Cargar datos reales de artists.csv
popularidades = []
seguidores = []

try:
    stats = pd.read_csv(os.path.join(BASE_FOLDER, "artists.csv"))
    popularidades = stats['popularity'].dropna().astype(int).tolist()
    seguidores = stats['followers'].dropna().astype(int).tolist()
except:
    print("No se encontraron o no se pudieron leer estadísticas detalladas.")

def calcular_estadisticas(valores):
    if not valores:
        return "N/A", "N/A", "N/A", "N/A"
    try:
        return round(mean(valores),2), median(valores), mode(valores), round(stdev(valores), 2)
    except StatisticsError:
        return round(mean(valores), 2), median(valores), "No única", round(stdev(valores), 2)
    
media_p, mediana_p, moda_p, desv_p = calcular_estadisticas(popularidades)
media_f, mediana_f, moda_f, desv_f = calcular_estadisticas(seguidores)

ws2.append(["Métrica", "Popularidad", "Seguidores"])
ws2.append(["Media", media_p, media_f])
ws2.append(["Mediana", mediana_p, mediana_f])
ws2.append(["Moda", moda_p, moda_f])
ws2.append(["Desviación estándar", desv_p, desv_f])

# 3.3 Hoja: Gráficas con imágenes
ws3 = wb.create_sheet("Gráficas")

for idx, path in enumerate([ruta1, ruta2, ruta3, ruta4]):
    img = XLImage(path)
    img.width = 480
    img.height = 300
    cell = f"A{idx * 20 + 1}"
    ws3.add_image(img, cell)

# Guardar archivo Excel
wb.save(OUTPUT_XLSX)
print(f"✓ Archivo generado: {OUTPUT_XLSX}")
